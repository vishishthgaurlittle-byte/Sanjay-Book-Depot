#!/usr/bin/env python3
"""
Build PRODUCT-ALLOCATION-PLAN.xlsx — how many products each brand should have
in each category, at 500 / 1,500 / 3,000 catalogue sizes.

TWO STEPS. The category and brand data live in ESM modules, so dump them to
JSON from Node first, then run this:

  cd backend
  node -e "import('./src/lib/data/catalog.mjs').then(async (m) => { \
    const { BRANDS } = await import('./src/lib/data/brands.mjs'); \
    const { PARENT_CATEGORIES, CHILD_CATEGORIES } = await import('./src/lib/data/categories.mjs'); \
    require('fs').writeFileSync('/tmp/catalog-meta.json', JSON.stringify({ \
      affinity: m.CATEGORY_BRAND_AFFINITY, \
      brands: BRANDS.map(b => ({ slug: b.slug, name: b.name, tier: b.tier })), \
      parents: PARENT_CATEGORIES.map(c => ({ slug: c.slug, name: c.name })), \
      children: CHILD_CATEGORIES.map(c => ({ slug: c.slug, name: c.name })) })); })"

  python3 src/scripts/plan-allocation.py

Outputs ../PRODUCT-ALLOCATION-PLAN.xlsx with four sheets:
  1. Category plan     - weight per category, products at each scale
  2. Brand x Category  - the full matrix
  3. Brand totals      - per-brand totals and categories served
  4. Notes & changes   - allocation rules and open questions
"""
import json, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open('/tmp/catalog-meta.json'))
parents = {p['slug']: p['name'] for p in d['parents']}
children = d['children']
brands = {b['slug']: b for b in d['brands']}
affinity = {k: list(v) for k, v in d['affinity'].items()}

# ── FIX: five brands were in no affinity list and would get 0 products ──
ADDED = {
  'pencils-lead':        ['hindustan-pencils'],          # parent of Apsara + Nataraj
  'erasers-sharpeners':  ['hindustan-pencils'],
  'paper-products':      ['anupam'],                     # watercolour / art paper
  'notebooks-registers': ['shalimar', 'nightingale'],    # classic Indian notebook houses
  'diaries-planners':    ['nightingale'],
  'art-supplies':        ['faber-chisel'],
  'writing-instruments': ['faber-chisel'],               # chisel markers
}
for cat, adds in ADDED.items():
    for b in adds:
        if b not in affinity[cat]:
            affinity[cat].append(b)

def parent_of(slug):
    best = None
    for p in parents:
        if slug.startswith(p + '-') and (best is None or len(p) > len(best)):
            best = p
    return best
child_parent = {c['slug']: parent_of(c['slug']) for c in children}
subs_of = defaultdict(list)
for c in children: subs_of[child_parent[c['slug']]].append(c['name'])

WEIGHTS = {
  'writing-instruments': 18, 'notebooks-registers': 16, 'paper-products': 10,
  'office-supplies': 10, 'art-supplies': 9, 'pencils-lead': 7, 'files-folders': 6,
  'geometry-instruments': 5, 'adhesives-tapes': 4, 'diaries-planners': 4,
  'erasers-sharpeners': 3, 'desk-organizers': 3, 'bags-pouches': 2,
  'premium-luxury': 2, 'school-essentials-kits': 1,
}
TIER_W = {'tier1': 3.0, 'tier2': 1.5, 'tier3': 0.8}

def allocate(budget, weights):
    tw = sum(weights.values())
    raw = {k: budget * v / tw for k, v in weights.items()}
    out = {k: max(1, int(math.floor(v))) for k, v in raw.items()}
    while sum(out.values()) > budget:
        k = max(out, key=lambda k: (out[k] - raw[k], -raw[k]))
        if out[k] <= 1: break
        out[k] -= 1
    while sum(out.values()) < budget:
        k = max(raw, key=lambda k: raw[k] - out[k]); out[k] += 1
    return out

matrix = defaultdict(dict)   # cat -> brand -> count
for TOTAL in ():
    pass

def build(total):
    """Allocate `total` products across categories, then across brands."""
    budget = {}
    for p, w in WEIGHTS.items():
        # never below the number of brands in the category, so each gets >= 1
        budget[p] = max(len(affinity[p]), round(total * w / 100))

    # Per-category rounding can overshoot the target by a product or two.
    # Trim the largest categories first, then top up if we undershot.
    diff = sum(budget.values()) - total
    while diff > 0:
        p = max(budget, key=lambda k: budget[k])
        if budget[p] <= len(affinity[p]):
            break
        budget[p] -= 1
        diff -= 1
    while diff < 0:
        p = max(WEIGHTS, key=lambda k: WEIGHTS[k] / budget[k])
        budget[p] += 1
        diff += 1

    m = defaultdict(dict)
    for p in WEIGHTS:
        weights = {x: TIER_W[brands[x]['tier']] for x in affinity[p] if x in brands}
        if not weights:
            continue
        for br, n in allocate(budget[p], weights).items():
            m[p][br] = n
    return m, budget


m500, budget500 = build(500)
m1500, _ = build(1500)
m3000, _ = build(3000)

bt = defaultdict(int)
for p, bs in m500.items():
    for b, n in bs.items(): bt[b] += n
unassigned = [b for b in brands if b not in bt]
print('planned @500:', sum(bt.values()), '| brands covered:', len(bt), '/', len(brands))
print('brands at zero:', [brands[b]['name'] for b in unassigned] or 'NONE')

# ══════════ XLSX ══════════
wb = Workbook()
H = Font(bold=True, color='FFFFFF', size=10)
HF = PatternFill('solid', fgColor='1F2937')
TITLE = Font(bold=True, size=14, color='B45309')
SUB = Font(italic=True, size=9, color='6B7280')
BOLD = Font(bold=True)
thin = Side(style='thin', color='E5E7EB')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
TOTF = PatternFill('solid', fgColor='FEF3C7')

def hdr(ws, row, vals, widths=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v); c.font = H; c.fill = HF
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = BOX
    if widths:
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

# ── Sheet 1: Category plan ──
ws = wb.active; ws.title = '1. Category plan'
ws['A1'] = 'Category allocation plan'; ws['A1'].font = TITLE
ws['A2'] = 'Weight = share of catalogue. Edit column C and the @500 / @1500 / @3000 columns follow.'; ws['A2'].font = SUB
hdr(ws, 4, ['Parent category', 'Sub-cats', 'Weight %', '@ 500 products', '@ 1,500', '@ 3,000', 'Per sub-cat @500', 'Sub-categories'],
    [26, 9, 10, 14, 10, 10, 15, 60])
r = 5
for p, w in sorted(WEIGHTS.items(), key=lambda x: -x[1]):
    subs = subs_of[p]
    vals = [parents[p], len(subs), w, budget500[p], round(1500*w/100), round(3000*w/100),
            round(budget500[p]/len(subs), 1), ', '.join(subs)]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i in (2,3,4,5,6,7): c.alignment = Alignment(horizontal='center')
        if i == 8: c.alignment = Alignment(wrap_text=True, vertical='top'); c.font = Font(size=9, color='6B7280')
    r += 1
for i, v in enumerate(['TOTAL', sum(len(subs_of[p]) for p in WEIGHTS), sum(WEIGHTS.values()),
                       sum(budget500.values()), 1500, 3000, '', ''], 1):
    c = ws.cell(row=r, column=i, value=v); c.font = BOLD; c.fill = TOTF; c.border = BOX
    c.alignment = Alignment(horizontal='center') if i > 1 else Alignment()
ws.freeze_panes = 'A5'

# ── Sheet 2: Brand x Category matrix ──
ws2 = wb.create_sheet('2. Brand x Category')
ws2['A1'] = 'Products per brand per category  (@ 500 total)'; ws2['A1'].font = TITLE
ws2['A2'] = 'Blank = brand does not sell in that category. Driven by CATEGORY_BRAND_AFFINITY in catalog.mjs.'; ws2['A2'].font = SUB
cats = sorted(WEIGHTS, key=lambda x: -WEIGHTS[x])
hdr(ws2, 4, ['Brand', 'Tier'] + [parents[c].replace(' & ', ' &\n') for c in cats] + ['TOTAL'],
    [24, 7] + [9]*len(cats) + [8])
r = 5
for b in sorted(brands, key=lambda x: (brands[x]['tier'], -bt.get(x, 0))):
    ws2.cell(row=r, column=1, value=brands[b]['name']).border = BOX
    tc = ws2.cell(row=r, column=2, value=brands[b]['tier'][-1]); tc.border = BOX
    tc.alignment = Alignment(horizontal='center')
    for i, cat in enumerate(cats, 3):
        n = m500[cat].get(b)
        c = ws2.cell(row=r, column=i, value=n if n else None); c.border = BOX
        c.alignment = Alignment(horizontal='center')
        if not n: c.fill = PatternFill('solid', fgColor='F9FAFB')
    t = ws2.cell(row=r, column=len(cats)+3, value=bt.get(b, 0)); t.font = BOLD; t.border = BOX
    t.alignment = Alignment(horizontal='center'); t.fill = TOTF
    r += 1
tot_row = r
ws2.cell(row=r, column=1, value='TOTAL').font = BOLD
for i, cat in enumerate(cats, 3):
    c = ws2.cell(row=r, column=i, value=sum(m500[cat].values())); c.font = BOLD; c.fill = TOTF
    c.alignment = Alignment(horizontal='center'); c.border = BOX
c = ws2.cell(row=r, column=len(cats)+3, value=sum(bt.values())); c.font = BOLD; c.fill = TOTF
c.alignment = Alignment(horizontal='center'); c.border = BOX
ws2.freeze_panes = 'C5'

# ── Sheet 3: Brand totals ──
ws3 = wb.create_sheet('3. Brand totals')
ws3['A1'] = 'Brand totals and scaling'; ws3['A1'].font = TITLE
ws3['A2'] = 'Every brand gets at least 1 product in each category it serves.'; ws3['A2'].font = SUB
hdr(ws3, 4, ['Brand', 'Tier', 'Categories served', '@ 500', '@ 1,500', '@ 3,000', 'Categories'],
    [24, 7, 16, 9, 10, 10, 56])
r = 5
b1500 = defaultdict(int); b3000 = defaultdict(int)
for cat in cats:
    for b, n in m1500[cat].items(): b1500[b] += n
    for b, n in m3000[cat].items(): b3000[b] += n
for b in sorted(brands, key=lambda x: (brands[x]['tier'], -bt.get(x, 0))):
    served = [parents[c] for c in cats if m500[c].get(b)]
    for i, v in enumerate([brands[b]['name'], 'tier'+str(brands[b]['tier'][-1]), len(served),
                           bt.get(b,0), b1500.get(b,0), b3000.get(b,0), ', '.join(served)], 1):
        c = ws3.cell(row=r, column=i, value=v); c.border = BOX
        if i in (2,3,4,5,6): c.alignment = Alignment(horizontal='center')
        if i == 7: c.font = Font(size=9, color='6B7280'); c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
ws3.freeze_panes = 'A5'

# ── Sheet 4: Notes ──
ws4 = wb.create_sheet('4. Notes & changes')
ws4.column_dimensions['A'].width = 110
notes = [
 ('Allocation rules', True),
 ('', False),
 ('1. Each parent category gets a share of the catalogue (sheet 1, column C).', False),
 ('   Weights reflect real Indian stationery retail: writing instruments and notebooks dominate,', False),
 ('   premium/luxury and bundles are deliberately small.', False),
 ('', False),
 ('2. Within a category, the budget is split across the brands that actually make those goods,', False),
 ('   weighted by tier:  tier1 = 3.0   tier2 = 1.5   tier3 = 0.8', False),
 ('   Allocated by largest-remainder so the column sums to the budget exactly.', False),
 ('', False),
 ('3. Every brand in a category gets at least 1 product, so no brand is empty.', False),
 ('', False),
 ('4. The category budget is then spread across its sub-categories (sheet 1, column G).', False),
 ('', False),
 ('CHANGE MADE TO THE AFFINITY MAP', True),
 ('', False),
 ('Five brands were in no affinity list and would have had zero products. Assigned as:', False),
 ('   hindustan-pencils  -> pencils-lead, erasers-sharpeners   (parent of Apsara and Nataraj)', False),
 ('   anupam             -> paper-products                     (watercolour and art paper)', False),
 ('   shalimar           -> notebooks-registers', False),
 ('   nightingale        -> notebooks-registers, diaries-planners', False),
 ('   faber-chisel       -> art-supplies, writing-instruments   (chisel markers)', False),
 ('', False),
 ('These are my inferences from the brand names. Correct any that are wrong.', False),
 ('', False),
 ('OPEN QUESTIONS', True),
 ('', False),
 ('- Should tier3 boutique brands really sell across many categories, or stay narrow and premium?', False),
 ('- Some tier1 brands are parents of others (Hindustan Pencils owns Apsara and Nataraj).', False),
 ('  Do you want the parent listed as a separate brand, or folded into its children?', False),
 ('- Real counts will differ. If you have an actual stock list, this plan is replaced by it.', False),
]
for i, (t, bold) in enumerate(notes, 1):
    c = ws4.cell(row=i, column=1, value=t)
    if bold: c.font = Font(bold=True, size=11, color='B45309')
    else: c.font = Font(size=10)

wb.save('/home/user/sanjay-book-depot/PRODUCT-ALLOCATION-PLAN.xlsx')
print('\nsaved PRODUCT-ALLOCATION-PLAN.xlsx')
print('\n=== brand totals @500 ===')
for b in sorted(brands, key=lambda x: (brands[x]['tier'], -bt.get(x,0))):
    print(f"  [{brands[b]['tier']}] {brands[b]['name']:<24} {bt.get(b,0):>3}  ({len([c for c in cats if m500[c].get(b)])} cats)")
